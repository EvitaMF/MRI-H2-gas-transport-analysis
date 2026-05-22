import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import os

# -------------------------------------------------
# Plot style
# -------------------------------------------------
plt.rcParams.update({
    "font.size": 12,
    "axes.labelsize": 12,
    "legend.fontsize": 9,
    "figure.dpi": 300
})
plt.rcParams["axes.spines.top"] = False
plt.rcParams["axes.spines.right"] = False

plot_kwargs = dict(
    marker='o',
    markersize=2,
    linestyle='--',
    linewidth=0.8
)

figsize_small = (5, 3)

# ============================================
# 1. SELECT DATASET
# ============================================

dataset = {
    "path": r"C:\Users\evita\OneDrive - University of Bergen\MRI_H2\diffusion_study\2026 02 Diffusion\MRI Scan data\CH4 N2 100 diff45\20260128_183601_CH4_CH4_N2_100_diff45_1_6\CH4_N2_100_diff45_2_diffusion_results.xlsx",
    "cutoff": 256695,
    "label": "CH$_4$–N$_2$ (100 bar)"
}

excel_path = dataset["path"]
TIME_CUTOFF = dataset["cutoff"]
LABEL = dataset["label"]

# ============================================
# 2. PHYSICAL PARAMETERS
# ============================================

V_A = 7.06858e-5
V_B = 7.06858e-5

A = 2.82743e-5
L = 0.201

# ============================================
# 3. READ DATA
# ============================================

df = pd.read_excel(excel_path)
df = df.sort_values("Time_seconds")

# ============================================
# VELG SLICE (HER!)
# ============================================

SELECTED_SLICE = 6   # 32 = axial, 6 = sagittal

df = df[df["Slice"] == SELECTED_SLICE]

# ============================================

t = df["Time_seconds"].values
S_A = df["ROI1_mean"].values

# ============================================
# 4. FILTER DATA
# ============================================

mask = t <= TIME_CUTOFF

t_fit_data = t[mask]
S_A_fit_data = S_A[mask]

print(f"\nUsing {len(t_fit_data)} points (t <= {TIME_CUTOFF} s)")

# ============================================
# 5. EXPONENTIAL MODEL
# ============================================

def exp_model(t, S_eq, S0, tau):
    return S_eq + (S0 - S_eq) * np.exp(-t / tau)

S_eq_guess = S_A_fit_data[-1]
S0_guess = S_A_fit_data[0]
tau_guess = (t_fit_data[-1] - t_fit_data[0]) / 5

popt, pcov = curve_fit(
    exp_model,
    t_fit_data,
    S_A_fit_data,
    p0=[S_eq_guess, S0_guess, tau_guess]
)

S_eq_fit, S0_fit, tau_fit = popt

# ============================================
# 6. UNCERTAINTY
# ============================================

tau_std = np.sqrt(pcov[2, 2])
D_dummy = 1 / (tau_fit * (1/V_A + 1/V_B))
D_dummy = D_dummy * L / A

# senere korrekt D brukes
# men vi trenger forholdet
# (eller bare bruk D direkte under)

# ============================================
# 7. RESULTS
# ============================================

k = 1 / (tau_fit * (1/V_A + 1/V_B))
D = k * L / A

# propagate uncertainty
D_std = D * (tau_std / tau_fit)

print("\n==============================")
print("FIT RESULTS")
print("==============================")
print(f"Tau = {tau_fit:.4f} ± {tau_std:.4f} s")
print(f"D = {D:.6e} ± {D_std:.2e} m^2/s")

# ============================================
# 8. PLOT
# ============================================

output_folder = os.path.dirname(excel_path)

t_hours = t / 3600
t_fit_data_hours = t_fit_data / 3600
cutoff_hours = TIME_CUTOFF / 3600

t_fit_full = np.linspace(t.min(), t.max(), 500)
t_fit_hours = t_fit_full / 3600
S_fit = exp_model(t_fit_full, S_eq_fit, S0_fit, tau_fit)

mask_unused = t > TIME_CUTOFF

plt.figure(figsize=figsize_small)

# used data
plt.plot(t_fit_data_hours, S_A_fit_data,
         'o--', color="tab:red", markersize=3,
         linewidth=0.8, label="CH$_4$ chamber")

# unused data
#plt.plot(t_hours[mask_unused], S_A[mask_unused],
#         'o', color="tab:red", alpha=0.15,
#         markersize=3, label="Excluded data")

# fit
plt.plot(t_fit_hours, S_fit,
         color="black", linewidth=1.2,
         label="Exponential fit")

# zoom
plt.xlim(0, t_fit_data_hours.max() * 1.05)

xmin, xmax = plt.gca().get_xlim()

# shaded region
#plt.axvspan(cutoff_hours, xmax,
#            color='lightgray',
#            alpha=0.35,
#            zorder=-1)

# y limits (include all data)
ymin_all = S_A.min()
ymax_all = S_A.max()

margin = 0.06 * (ymax_all - ymin_all)
plt.ylim(ymin_all - margin, ymax_all + margin)

plt.xlabel("Time [hours]")
plt.ylabel("Signal intensity (a.u.)")
plt.grid(alpha=0.3)
plt.legend(frameon=False)

plt.tight_layout()

save_path = os.path.join(output_folder, "diffusion_fit_sag.png")
plt.savefig(save_path)

plt.show()

print("Plot saved to:", save_path)

# ============================================
# 9. SAVE TO EXCEL (APPEND ROWS)
# ============================================

from openpyxl import load_workbook

results_dict = {
    "Dataset": LABEL,
    "Slice": SELECTED_SLICE,

    "Time_cutoff_s": TIME_CUTOFF,

    "Tau_s": tau_fit,
    "Tau_std_s": tau_std,

    "D_m2_per_s": D,
    "D_std_m2_per_s": D_std,

    "D_cm2_per_s": D * 1e4,
    "D_std_cm2_per_s": D_std * 1e4,

    "D_formatted_cm2_per_s": f"{D*1e4:.3f} ± {D_std*1e4:.3f}",

    "N_points_used": len(t_fit_data)
}

results_df = pd.DataFrame([results_dict])

sheet_name = "Fit_results"

# hvis arket finnes → append
if sheet_name in load_workbook(excel_path).sheetnames:
    existing_df = pd.read_excel(excel_path, sheet_name=sheet_name)
    results_df = pd.concat([existing_df, results_df], ignore_index=True)

# skriv tilbake
with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    results_df.to_excel(writer, sheet_name=sheet_name, index=False)

print("Results updated in Excel (appended row)")